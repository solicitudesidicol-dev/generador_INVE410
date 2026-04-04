import sys
import math
import random
import pandas as pd
from pathlib import Path
from PyQt6.QtWidgets import (QApplication, QMainWindow, QVBoxLayout, QHBoxLayout,
                             QWidget, QPushButton, QTableWidget, QTableWidgetItem,
                             QHeaderView, QLabel, QMessageBox, QStatusBar, QGroupBox,
                             QStyledItemDelegate, QStyle)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QFont, QColor, QPalette


# ─────────────────────────────────────────────────────────────────────────────
# Delegate: respeta setForeground incluso en filas seleccionadas
# ─────────────────────────────────────────────────────────────────────────────
class ColorDelegate(QStyledItemDelegate):
    def paint(self, painter, option, index):
        fg = index.data(Qt.ItemDataRole.ForegroundRole)
        opt = option.__class__(option)
        self.initStyleOption(opt, index)

        if fg is not None:
            try:
                text_color = fg.color() if hasattr(fg, "color") else QColor(fg)
            except Exception:
                text_color = QColor("#2c3e50")
        else:
            if opt.state & QStyle.StateFlag.State_Selected:
                text_color = QColor("#ffffff")
            else:
                text_color = QColor("#2c3e50")

        opt.text = ""
        style = opt.widget.style() if opt.widget else QApplication.style()
        style.drawControl(QStyle.ControlElement.CE_ItemViewItem, opt, painter, opt.widget)

        painter.save()
        painter.setPen(text_color)
        text_rect = style.subElementRect(
            QStyle.SubElement.SE_ItemViewItemText, opt, opt.widget
        )
        painter.drawText(
            text_rect,
            int(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft),
            index.data(Qt.ItemDataRole.DisplayRole) or ""
        )
        painter.restore()


# ─────────────────────────────────────────────────────────────────────────────
# Parámetros de generación por diámetro nominal
# ─────────────────────────────────────────────────────────────────────────────
PARAMS = {
    101: {
        "longitud_nominal": 201.0,
        "masa_min": 3300,
        "masa_max": 3600,
    },
    151: {
        "longitud_nominal": 301.0,
        "masa_min": 12000,
        "masa_max": 13500,
    },
}
VARIACION = 0.015   # ±1.5% para diámetros y longitudes


def _val(nominal, variacion=VARIACION):
    """Valor aleatorio con variación ±variacion alrededor del nominal."""
    return round(nominal * (1 + random.uniform(-variacion, variacion)), 2)


def _generar_submuestra(diametro_nominal, resistencia_nominal, esfuerzo_mpa_deseado):
    """
    Genera una submuestra completa (diámetros, longitudes, masa, carga)
    para un diámetro nominal dado y un esfuerzo MPa objetivo.
    Devuelve un dict con todas las claves necesarias.
    """
    p = PARAMS[diametro_nominal]

    d1 = _val(diametro_nominal)
    d2 = _val(diametro_nominal)
    l1 = _val(p["longitud_nominal"])
    l2 = _val(p["longitud_nominal"])
    l3 = _val(p["longitud_nominal"])

    masa = random.randint(p["masa_min"], p["masa_max"])

    # Carga a partir del esfuerzo deseado y el diámetro promedio
    d_prom = (d1 + d2) / 2
    area = math.pi * (d_prom ** 2) / 4   # mm²
    carga = round((esfuerzo_mpa_deseado * area) / 1000, 1)   # kg

    return {
        "d1": d1, "d2": d2,
        "l1": l1, "l2": l2, "l3": l3,
        "masa": masa,
        "carga": carga,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Ventana principal de Pendientes
# ─────────────────────────────────────────────────────────────────────────────
class PendientesWindow(QMainWindow):
    def __init__(self, excel_path: Path):
        super().__init__()
        self.excel_path = excel_path
        self.df = None
        self.visible_columns = []

        self.setWindowTitle("📋 Pendientes por Generar")
        self.setGeometry(100, 100, 1900, 900)

        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)

        # HEADER
        header = QLabel("📋 PENDIENTES - GENERADOR DE DATOS")
        header.setFont(QFont("Arial", 16, weight=75))
        header.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header.setStyleSheet("color:#000000; padding:12px; font-weight:bold;")
        header.setFixedHeight(50)
        main_layout.addWidget(header)

        # CONTROLES
        controls = QHBoxLayout()

        self.btn_reload = QPushButton("🔄 Recargar")
        self.btn_reload.setObjectName("btn_reload")
        self.btn_reload.clicked.connect(self.load_data)

        self.lbl_row = QLabel("Fila: -")
        self.lbl_row.setStyleSheet("font-size:14px; font-weight:bold; padding:8px;")

        self.lbl_status = QLabel("🟢 Listo")
        self._set_status("ok", "🟢 Listo")

        self.btn_save = QPushButton("💾 GUARDAR")
        self.btn_save.setObjectName("btn_save")
        self.btn_save.clicked.connect(self.save_all)

        controls.addWidget(self.btn_reload)
        controls.addStretch()
        controls.addWidget(self.lbl_row)
        controls.addWidget(self.lbl_status)
        controls.addStretch()
        controls.addWidget(self.btn_save)
        main_layout.addLayout(controls)

        # TABLA
        group = QGroupBox("📊 PENDIENTES - Edita 'Esfuerzo MPa Promedio' para generar datos")
        table_layout = QVBoxLayout(group)
        self.table = QTableWidget()
        self.table.setItemDelegate(ColorDelegate(self.table))
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.itemChanged.connect(self.on_cell_changed)
        self.table.itemClicked.connect(self.on_row_selected)
        table_layout.addWidget(self.table)
        main_layout.addWidget(group)

        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)

        self.is_updating = False
        self.load_data()

    # ──────────────────────────────────────────────────────────────────────
    # Estado visual
    # ──────────────────────────────────────────────────────────────────────
    def _set_status(self, estado: str, texto: str):
        colores = {
            "ok":      ("color:#27ae60;", "background:rgba(39,174,96,0.15);"),
            "error":   ("color:#e74c3c;", "background:rgba(231,76,60,0.15);"),
            "loading": ("color:#f39c12;", "background:rgba(243,156,18,0.15);"),
        }
        c, bg = colores.get(estado, colores["ok"])
        self.lbl_status.setStyleSheet(
            f"font-size:13px; padding:7px; border-radius:6px; {c} {bg}"
        )
        self.lbl_status.setText(texto)

    # ──────────────────────────────────────────────────────────────────────
    # Carga de datos
    # ──────────────────────────────────────────────────────────────────────
    def load_data(self):
        try:
            self._set_status("loading", "⏳ Cargando...")
            QApplication.processEvents()
            self.df = pd.read_excel(self.excel_path, sheet_name="Pendientes_Generar")

            # Normalizar columna Diámetro (puede venir como int o float)
            if "Diámetro" in self.df.columns:
                self.df["Diámetro"] = pd.to_numeric(
                    self.df["Diámetro"], errors="coerce"
                ).astype("Int64")

            # Normalizar Cantidad
            if "Cantidad" in self.df.columns:
                self.df["Cantidad"] = pd.to_numeric(
                    self.df["Cantidad"], errors="coerce"
                ).fillna(1).astype(int)

            # Agregar columnas de medición si no existen
            cols_medicion = [
                "Esfuerzo MPa Promedio",
                "Diámetro 1-1", "Diámetro 1-2",
                "Longitud 1-1", "Longitud 1-2", "Longitud 1-3",
                "Masa 1", "Carga 1",
                "Diámetro 2-1", "Diámetro 2-2",
                "Longitud 2-1", "Longitud 2-2", "Longitud 2-3",
                "Masa 2", "Carga 2",
                "Diámetro 3-1", "Diámetro 3-2",
                "Longitud 3-1", "Longitud 3-2", "Longitud 3-3",
                "Masa 3", "Carga 3",
            ]
            for col in cols_medicion:
                if col not in self.df.columns:
                    self.df[col] = None

            self.populate_table()
            self._set_status("ok", f"✅ {len(self.df)} pendientes cargados")
            self.status_bar.showMessage(f"✅ {len(self.df)} filas")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
            self._set_status("error", f"❌ {e}")

    # ──────────────────────────────────────────────────────────────────────
    # Poblar tabla
    # ──────────────────────────────────────────────────────────────────────
    def populate_table(self):
        if self.df is None:
            return

        # Columnas a mostrar — primero las informativas, luego las de medición
        cols_info = [
            "consecutivo", "idext", "Resistencia nominal", "codobraconf",
            "Diámetro", "Cantidad",
        ]
        cols_medicion = [
            "Esfuerzo MPa Promedio",
            "Diámetro 1-1", "Diámetro 1-2",
            "Longitud 1-1", "Longitud 1-2", "Longitud 1-3",
            "Masa 1", "Carga 1",
            "Diámetro 2-1", "Diámetro 2-2",
            "Longitud 2-1", "Longitud 2-2", "Longitud 2-3",
            "Masa 2", "Carga 2",
            "Diámetro 3-1", "Diámetro 3-2",
            "Longitud 3-1", "Longitud 3-2", "Longitud 3-3",
            "Masa 3", "Carga 3",
        ]

        self.visible_columns = [
            c for c in cols_info + cols_medicion if c in self.df.columns
        ]

        # Columnas de solo lectura (informativas)
        self.readonly_columns = set(cols_info)
        # Columna objetivo del optimizador
        self.target_column = "Esfuerzo MPa Promedio"
        # Columnas generadas automáticamente (no editables salvo Esfuerzo MPa Promedio)
        self.generated_columns = set(cols_medicion) - {self.target_column}

        self.table.itemChanged.disconnect(self.on_cell_changed)
        self.table.setRowCount(len(self.df))
        self.table.setColumnCount(len(self.visible_columns))
        self.table.setHorizontalHeaderLabels(self.visible_columns)

        for i in range(len(self.df)):
            for j, col in enumerate(self.visible_columns):
                val = self.df.iloc[i][col]
                display = self._fmt(col, val)
                item = QTableWidgetItem(display)
                item.setForeground(QColor("#2c3e50"))

                if col in self.readonly_columns:
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    item.setBackground(Qt.GlobalColor.lightGray)
                    item.setToolTip("🔒 Solo lectura")
                elif col == self.target_column:
                    item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                    item.setToolTip("🎯 Ingresa el esfuerzo MPa objetivo para generar datos")
                else:
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    item.setBackground(QColor("#eaf4fb"))
                    item.setToolTip("⚙️ Generado automáticamente")

                self.table.setItem(i, j, item)

        self.table.itemChanged.connect(self.on_cell_changed)
        self.table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive
        )
        self.table.resizeColumnsToContents()
        self.table.selectRow(0)

    def _fmt(self, col, val):
        """Formatea un valor para mostrar en tabla."""
        try:
            if pd.isna(val) or val is None or val == "":
                return ""
        except Exception:
            pass
        if any(x in col for x in ["Diámetro", "Longitud", "Esfuerzo"]):
            try:
                return f"{float(val):.2f}"
            except Exception:
                pass
        return str(val)

    # ──────────────────────────────────────────────────────────────────────
    # Eventos de tabla
    # ──────────────────────────────────────────────────────────────────────
    def on_row_selected(self, item=None):
        row = self.table.currentRow()
        if row >= 0:
            self.lbl_row.setText(f"Fila: {row + 1}")

    def on_cell_changed(self, item):
        if self.is_updating or self.df is None:
            return

        row = item.row()
        col_idx = item.column()
        col_name = self.visible_columns[col_idx]

        # Solo reaccionar al cambio de la columna objetivo
        if col_name != self.target_column:
            return

        try:
            esfuerzo = float(item.text().strip())
            if esfuerzo <= 0:
                raise ValueError
        except (ValueError, AttributeError):
            return

        self._generar_fila(row, esfuerzo)

    # ──────────────────────────────────────────────────────────────────────
    # Generación de datos
    # ──────────────────────────────────────────────────────────────────────
    def _generar_fila(self, row: int, esfuerzo_mpa: float):
        """Genera datos de medición para una fila usando el esfuerzo MPa objetivo."""
        try:
            diametro_nominal = int(self.df.iloc[row]["Diámetro"])
            cantidad = int(self.df.iloc[row]["Cantidad"])
        except (ValueError, TypeError, KeyError):
            self._set_status("error", f"❌ Fila {row+1}: diámetro o cantidad inválidos")
            return

        if diametro_nominal not in PARAMS:
            self._set_status(
                "error",
                f"❌ Diámetro {diametro_nominal} no soportado (use 101 o 151)"
            )
            return

        # Limitar submuestras a máximo 3
        n = min(cantidad, 3)

        # Generar submuestras
        submuestras = [
            _generar_submuestra(diametro_nominal, None, esfuerzo_mpa)
            for _ in range(n)
        ]

        # Mapeo submuestra → columnas del DataFrame
        campos = {
            1: ("Diámetro 1-1","Diámetro 1-2","Longitud 1-1","Longitud 1-2","Longitud 1-3","Masa 1","Carga 1"),
            2: ("Diámetro 2-1","Diámetro 2-2","Longitud 2-1","Longitud 2-2","Longitud 2-3","Masa 2","Carga 2"),
            3: ("Diámetro 3-1","Diámetro 3-2","Longitud 3-1","Longitud 3-2","Longitud 3-3","Masa 3","Carga 3"),
        }

        self.is_updating = True
        self.table.itemChanged.disconnect(self.on_cell_changed)
        try:
            for idx, sm in enumerate(submuestras, start=1):
                cols = campos[idx]
                valores = [
                    sm["d1"], sm["d2"],
                    sm["l1"], sm["l2"], sm["l3"],
                    sm["masa"], sm["carga"],
                ]
                for col, val in zip(cols, valores):
                    self.df.at[row, col] = val
                    if col in self.visible_columns:
                        j = self.visible_columns.index(col)
                        item = self.table.item(row, j)
                        if item:
                            item.setText(self._fmt(col, val))

            # Guardar esfuerzo en df
            self.df.at[row, self.target_column] = esfuerzo_mpa

            self._set_status(
                "ok",
                f"✅ Fila {row+1}: {n} submuestras generadas para {esfuerzo_mpa:.1f} MPa"
            )
            self.status_bar.showMessage(
                f"Fila {row+1} generada — {n} submuestras — diámetro {diametro_nominal} mm"
            )
        except Exception as e:
            self._set_status("error", f"❌ Error generando fila {row+1}: {e}")
        finally:
            self.is_updating = False
            self.table.itemChanged.connect(self.on_cell_changed)

    # ──────────────────────────────────────────────────────────────────────
    # Guardado
    # ──────────────────────────────────────────────────────────────────────
    def save_all(self):
        if self.df is None:
            self._set_status("error", "❌ No hay datos cargados")
            return

        try:
            from openpyxl import load_workbook, Workbook

            self._set_status("loading", "💾 Guardando...")
            QApplication.processEvents()

            output_path = self.excel_path.parent / "pendientes_final.xlsx"

            if output_path.exists():
                try:
                    wb = load_workbook(output_path)
                except Exception:
                    wb = Workbook()
            else:
                wb = Workbook()

            sheet_name = "Pendientes_final"
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            ws = wb.create_sheet(sheet_name)

            # Headers
            for c_num, col_name in enumerate(self.df.columns, 1):
                ws.cell(row=1, column=c_num, value=str(col_name))

            # Datos con tipos conservados
            for r_idx, (_, row_data) in enumerate(self.df.iterrows()):
                r_num = r_idx + 2
                for c_num, value in enumerate(row_data.values, 1):
                    if pd.isna(value) if not isinstance(value, str) else False:
                        cell_value = None
                    else:
                        try:
                            fv = float(value)
                            cell_value = int(fv) if fv == int(fv) else fv
                        except (ValueError, TypeError):
                            cell_value = str(value)
                    ws.cell(row=r_num, column=c_num, value=cell_value)

            wb.save(output_path)

            filas = len(self.df)
            self._set_status("ok", f"💾 Guardado ✓ ({filas} filas)")
            self.status_bar.showMessage(f"✅ {sheet_name} guardado — {filas} filas")

        except Exception as e:
            self._set_status("error", f"❌ Error al guardar: {e}")
            print(f"❌ save_all pendientes: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# Función para abrir la ventana desde app.py
# ─────────────────────────────────────────────────────────────────────────────
def abrir_pendientes(excel_path: Path, parent=None):
    """Abre la ventana de Pendientes. Llamar desde app.py."""
    win = PendientesWindow(excel_path)
    win.show()
    return win  # guardar referencia para que no sea recolectado por GC


# ─────────────────────────────────────────────────────────────────────────────
# Estilos (idénticos a app.py para coherencia visual)
# ─────────────────────────────────────────────────────────────────────────────
STYLESHEET = """
    QMainWindow {
        background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
            stop:0 #2c3e50, stop:1 #34495e);
    }
    QGroupBox {
        font-weight: bold;
        font-size: 13px;
        border: 1px solid #27ae60;
        border-radius: 10px;
        margin: 12px;
        padding-top: 18px;
        background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
            stop:0 #ffffff, stop:1 #f8f9fa);
        color: #2c3e50;
    }
    QGroupBox::title {
        subcontrol-origin: margin;
        left: 16px;
        padding: 0 10px 4px 10px;
        color: white;
        background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
            stop:0 #27ae60, stop:1 #2ecc71);
        border-radius: 7px;
        font-size: 12px;
    }
    QTableWidget {
        gridline-color: #dee2e6;
        background: white;
        font-size: 11px;
        font-family: 'Segoe UI', Arial;
        border: 1px solid #2c3e50;
        alternate-background-color: #f8f9fa;
    }
    QTableWidget::item {
        padding: 1px;
        border-bottom: 1px solid #e9ecef;
    }
    QTableWidget::item:selected {
        background: #27ae60;
        font-weight: bold;
    }
    QTableWidget::item:focus {
        border: 2px solid #1e8449;
        background: rgba(30,132,73,0.25);
    }
    QHeaderView::section {
        background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
            stop:0 #2c3e50, stop:1 #34495e);
        color: #ffffff;
        font-weight: bold;
        padding: 10px;
        border: none;
        font-size: 11px;
    }
    QPushButton {
        padding: 12px 24px;
        font-weight: bold;
        border-radius: 8px;
        border: 2px solid transparent;
        font-size: 13px;
        font-family: 'Segoe UI', Arial;
        min-width: 120px;
    }
    QPushButton#btn_save {
        background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
            stop:0 #2ecc71, stop:1 #58d68d);
        color: white;
        border-color: #27ae60;
    }
    QPushButton#btn_save:hover {
        background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
            stop:0 #27ae60, stop:1 #2ecc71);
    }
    QPushButton#btn_reload {
        background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
            stop:0 #95a5a6, stop:1 #bdc3c7);
        color: #2c3e50;
        border-color: #7f8c8d;
    }
    QPushButton#btn_reload:hover {
        background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
            stop:0 #7f8c8d, stop:1 #95a5a6);
    }
    QLabel {
        color: #2c3e50;
        font-weight: 600;
        padding: 6px;
        background: rgba(255,255,255,0.8);
        border-radius: 5px;
    }
    QStatusBar {
        background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
            stop:0 #34495e, stop:1 #2c3e50);
        color: #ecf0f1;
        font-weight: 500;
        padding: 6px;
        border-top: 2px solid #27ae60;
    }
    QScrollBar:vertical {
        background: #ecf0f1;
        width: 12px;
        border-radius: 6px;
    }
    QScrollBar::handle:vertical {
        background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
            stop:0 #27ae60, stop:1 #2ecc71);
        border-radius: 6px;
        min-height: 20px;
    }
"""


# ─────────────────────────────────────────────────────────────────────────────
# Entry point independiente (para pruebas directas)
# ─────────────────────────────────────────────────────────────────────────────
def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    palette = app.palette()
    palette.setColor(QPalette.ColorRole.Highlight,       QColor("#1e8449"))
    palette.setColor(QPalette.ColorRole.HighlightedText, QColor("#ffffff"))
    app.setPalette(palette)

    app.setStyleSheet(STYLESHEET)

    excel_path = Path("..") / "data" / "prueba.xlsx"
    if not excel_path.exists():
        QMessageBox.critical(None, "Error", f"No se encontró:\n{excel_path}")
        sys.exit(1)

    win = PendientesWindow(excel_path)
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()