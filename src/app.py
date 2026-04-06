import sys
import pandas as pd
import os
import math
import random
from pathlib import Path
from PyQt6.QtWidgets import (QApplication, QMainWindow, QVBoxLayout, QHBoxLayout,
                             QWidget, QPushButton, QTableWidget, QTableWidgetItem,
                             QHeaderView, QLabel, QMessageBox, QStatusBar, QGroupBox,
                             QStyledItemDelegate, QStyle, QTabWidget)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QFont, QColor, QPalette



class ColorDelegate(QStyledItemDelegate):
    """
    Delegate que respeta el color de foreground del item (rol ForegroundRole)
    incluso cuando la fila está seleccionada. El fondo de selección se mantiene.
    """
    def paint(self, painter, option, index):
        # Obtener el color asignado al item via setForeground
        fg = index.data(Qt.ItemDataRole.ForegroundRole)

        # Dibujar fondo y decoraciones normales (selección, foco, etc.)
        # pero sin que Qt sobreescriba el color de texto
        opt = option.__class__(option)
        self.initStyleOption(opt, index)

        # ForegroundRole devuelve un QBrush, no un QColor — extraer el color
        # si no, usar blanco en selección y oscuro en el resto
        if fg is not None:
            try:
                # QBrush.color() devuelve el QColor
                text_color = fg.color() if hasattr(fg, "color") else QColor(fg)
            except Exception:
                text_color = QColor("#2c3e50")
        else:
            if opt.state & QStyle.StateFlag.State_Selected:
                text_color = QColor("#ffffff")
            else:
                text_color = QColor("#2c3e50")

        # Quitar el texto del option para que la clase base no lo pinte
        opt.text = ""
        style = opt.widget.style() if opt.widget else QApplication.style()
        style.drawControl(QStyle.ControlElement.CE_ItemViewItem, opt, painter, opt.widget)

        # Pintar el texto manualmente con el color correcto
        painter.save()
        painter.setPen(text_color)
        text_rect = style.subElementRect(
            QStyle.SubElement.SE_ItemViewItemText, opt, opt.widget
        )
        painter.drawText(
            text_rect,
            int(Qt.AlignmentFlag.AlignVCenter),
            index.data(Qt.ItemDataRole.DisplayRole) or ""
        )
        painter.restore()


class ExcelDynamicApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.df = None
        self.current_file = ""
        self.selected_row = 0
        self.is_updating = False
        self._undo_stack = []  # lista de (row, col_name, valor_anterior, valor_nuevo)
        self._redo_stack = []
        self._max_undo = 50
        self.change_timer = QTimer()
        self.change_timer.setSingleShot(True)
        self.change_timer.timeout.connect(self._full_recalculate_row)
        self.init_ui()
        self.auto_load_excel()

    def init_ui(self):
        self.setWindowTitle("🎯 INV E 410 Pro - OPTIMIZADOR DE DATOS INTELIGENTE")
        self.setGeometry(50, 50, 1900, 950)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # HEADER
        header = QLabel("EDITOR + OPTIMIZADOR INTELIGENTE DE DATOS")
        header.setFont(QFont("Arial", 18, weight=75))
        header.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header.setStyleSheet("color:#000000; padding:15px; font-weight:bold;")
        header.setFixedHeight(60)
        main_layout.addWidget(header)

        # ── TAB WIDGET ────────────────────────────────────────────────────
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #27ae60;
                border-radius: 8px;
                background: white;
            }
            QTabBar::tab {
                background: #34495e;
                color: #ecf0f1;
                padding: 10px 28px;
                font-size: 13px;
                font-weight: bold;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                margin-right: 3px;
            }
            QTabBar::tab:selected {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #27ae60, stop:1 #2ecc71);
                color: white;
            }
            QTabBar::tab:hover:!selected { background: #4a6278; }
        """)
        self.tabs.currentChanged.connect(self._on_tab_changed)
        main_layout.addWidget(self.tabs)

        # ── PESTAÑA 1: CON DATOS ──────────────────────────────────────────
        tab1 = QWidget()
        tab1_layout = QVBoxLayout(tab1)

        # Controles pestaña 1
        ctrl1 = QHBoxLayout()
        self.btn_reload = QPushButton("🔄 Recargar")
        self.btn_reload.setObjectName("btn_reload")
        self.btn_reload.clicked.connect(self.auto_load_excel)
        self.btn_save = QPushButton("💾 GUARDAR")
        self.btn_save.setObjectName("btn_save")
        self.btn_save.clicked.connect(self.save_all)
        self.lbl_row = QLabel("Fila: -")
        self.lbl_row.setStyleSheet("font-size:15px; font-weight:bold; padding:8px;")
        self.lbl_status = QLabel("🟢 Listo")
        self._set_status("ok", "🟢 Listo")

        ctrl1.addWidget(self.btn_reload)
        ctrl1.addStretch()
        ctrl1.addWidget(self.lbl_row)
        ctrl1.addWidget(self.lbl_status)
        ctrl1.addStretch()
        ctrl1.addWidget(self.btn_save)
        tab1_layout.addLayout(ctrl1)

        # Tabla pestaña 1
        group1 = QGroupBox("📊 TABLA EDITABLE - Doble clic para editar")
        g1_layout = QVBoxLayout(group1)
        self.table = QTableWidget()
        self.table.setItemDelegate(ColorDelegate(self.table))
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.itemChanged.connect(self.on_cell_changed)
        self.table.itemClicked.connect(self.on_row_selected)
        g1_layout.addWidget(self.table)
        tab1_layout.addWidget(group1)
        self.tabs.addTab(tab1, "📊 Con Datos")

        # ── PESTAÑA 2: PENDIENTES ─────────────────────────────────────────
        tab2 = QWidget()
        tab2_layout = QVBoxLayout(tab2)

        # Controles pestaña 2
        ctrl2 = QHBoxLayout()
        self.btn_reload_p = QPushButton("🔄 Recargar")
        self.btn_reload_p.setObjectName("btn_reload")
        self.btn_reload_p.clicked.connect(self.load_pendientes)
        self.btn_save_p = QPushButton("💾 GUARDAR")
        self.btn_save_p.setObjectName("btn_save")
        self.btn_save_p.clicked.connect(self.save_pendientes)
        self.lbl_row_p = QLabel("Fila: -")
        self.lbl_row_p.setStyleSheet("font-size:15px; font-weight:bold; padding:8px;")
        self.lbl_status_p = QLabel("🟢 Listo")
        self._set_status_p("ok", "🟢 Listo")

        ctrl2.addWidget(self.btn_reload_p)
        ctrl2.addStretch()
        ctrl2.addWidget(self.lbl_row_p)
        ctrl2.addWidget(self.lbl_status_p)
        ctrl2.addStretch()
        ctrl2.addWidget(self.btn_save_p)
        tab2_layout.addLayout(ctrl2)

        # Tabla pestaña 2
        group2 = QGroupBox("📋 PENDIENTES - Edita 'Esfuerzo MPa Promedio' para generar datos")
        g2_layout = QVBoxLayout(group2)
        self.table_p = QTableWidget()
        self.table_p.setItemDelegate(ColorDelegate(self.table_p))
        self.table_p.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table_p.itemChanged.connect(self.on_cell_changed_p)
        self.table_p.itemClicked.connect(self.on_row_selected_p)
        g2_layout.addWidget(self.table_p)
        tab2_layout.addWidget(group2)
        self.tabs.addTab(tab2, "📋 Pendientes")

        # ── STATUS BAR ────────────────────────────────────────────────────
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)

        # Atajos Ctrl+Z / Ctrl+Y
        self.is_updating_p = False
        self.df_p = None
        self.visible_columns_p = []


    # ─────────────────────────────────────────────
    # FIX #2: helper centralizado para cambiar
    # estado del label y reaplicar el stylesheet
    # ─────────────────────────────────────────────
    def _set_status(self, estado: str, texto: str):
        """
        estado: 'ok' | 'error' | 'loading'
        Actualiza texto y reaplica el stylesheet dinámico correctamente.
        """
        colores = {
            "ok":      ("color: #27ae60;", "background: rgba(39,174,96,0.15);"),
            "error":   ("color: #e74c3c;", "background: rgba(231,76,60,0.15);"),
            "loading": ("color: #f39c12;", "background: rgba(243,156,18,0.15);"),
        }
        color, bg = colores.get(estado, colores["ok"])
        self.lbl_status.setStyleSheet(
            f"font-size:14px; padding:8px; border-radius:6px; {color} {bg}"
        )
        self.lbl_status.setText(texto)

    # ─────────────────────────────────────────────
    # Carga de datos
    # ─────────────────────────────────────────────
    def auto_load_excel(self):
        data_path = Path(r"\\dc01\JEFELAB\DOCUMENTOS\1. Informes de ensayo de laboratorio\Informes de concreto")
        input_file = data_path / "verificacion_2.xlsm"
        try:
            data_path.mkdir(exist_ok=True)
            if input_file.exists():
                
                # 🔥 NUEVO: REFRESH ANTES DE CARGAR
                self._set_status("loading", "🔄 Actualizando Excel...")
                QApplication.processEvents()

                self.refresh_excel(input_file)
                
                self.status_bar.showMessage("⏳ Cargando...")
                QApplication.processEvents()
                self.df = pd.read_excel(input_file, sheet_name="INV E 410")
                self.current_file = str(input_file)

                # Recalcular EDAD en Python (date2 - date1 en días)
                # Equivale a la fórmula Excel: =SI(J2=0,"-", J2-I2)
                if "date1" in self.df.columns and "date2" in self.df.columns:
                    d1 = pd.to_datetime(self.df["date1"], errors="coerce")
                    d2 = pd.to_datetime(self.df["date2"], errors="coerce")
                    # Resultado: entero de días; NaN si alguna fecha falta
                    self.df["EDAD"] = (d2 - d1).dt.days

                self.populate_table()
                self.status_bar.showMessage(f"✅ {len(self.df)} filas cargadas")
            else:
                QMessageBox.warning(self, "Falta archivo",
                                    f"Copia 'prueba.xlsx' a:\n{input_file}")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    # ─────────────────────────────────────────────
    # FIX #6: bloquear señal itemChanged durante
    # la carga inicial de la tabla
    # ─────────────────────────────────────────────
    def populate_table(self):
        if self.df is None:
            return

        columnas_mostrar = [
            "consecutivo", "idext", "Resistencia nominal", "codobraconf", "EDAD",
            "Esfuerzo MPa Promedio", "Esfuerzo % Promedio",
            "Densidad M1", "Diámetro M1", "Esfuerzo MPa M1", "Esfuerzo % M1",
            "Densidad M2", "Diámetro M2", "Esfuerzo MPa M2", "Esfuerzo % M2",
            "Densidad M3", "Diámetro M3", "Esfuerzo MPa M3", "Esfuerzo % M3",
            "Diámetro 1-1", "Diámetro 1-2", "Longitud 1-1", "Longitud 1-2", "Longitud 1-3",
            "Masa 1", "Carga 1",
            "Diámetro 2-1", "Diámetro 2-2", "Longitud 2-1", "Longitud 2-2", "Longitud 2-3",
            "Masa 2", "Carga 2",
            "Diámetro 3-1", "Diámetro 3-2", "Longitud 3-1", "Longitud 3-2", "Longitud 3-3",
            "Masa 3", "Carga 3",
        ]

        self.visible_columns = [c for c in columnas_mostrar if c in self.df.columns]

        self.calculated_columns = {
            "Diámetro M1": lambda i: self.promedio(i, ["Diámetro 1-1", "Diámetro 1-2"]),
            "Diámetro M2": lambda i: self.promedio(i, ["Diámetro 2-1", "Diámetro 2-2"]),
            "Diámetro M3": lambda i: self.promedio(i, ["Diámetro 3-1", "Diámetro 3-2"]),
            "Densidad M1": lambda i: self.densidad(
                i, "Masa 1", ["Diámetro 1-1", "Diámetro 1-2"],
                ["Longitud 1-1", "Longitud 1-2", "Longitud 1-3"]),
            "Densidad M2": lambda i: self.densidad(
                i, "Masa 2", ["Diámetro 2-1", "Diámetro 2-2"],
                ["Longitud 2-1", "Longitud 2-2", "Longitud 2-3"]),
            "Densidad M3": lambda i: self.densidad(
                i, "Masa 3", ["Diámetro 3-1", "Diámetro 3-2"],
                ["Longitud 3-1", "Longitud 3-2", "Longitud 3-3"]),
            "Esfuerzo MPa M1": lambda i: self.esfuerzo(
                i, "Carga 1", ["Diámetro 1-1", "Diámetro 1-2"]),
            "Esfuerzo MPa M2": lambda i: self.esfuerzo(
                i, "Carga 2", ["Diámetro 2-1", "Diámetro 2-2"]),
            "Esfuerzo MPa M3": lambda i: self.esfuerzo(
                i, "Carga 3", ["Diámetro 3-1", "Diámetro 3-2"]),
            "Esfuerzo % M1": lambda i: self.esfuerzo_pct_calc(
                i, "Esfuerzo MPa M1", "Resistencia nominal"),
            "Esfuerzo % M2": lambda i: self.esfuerzo_pct_calc(
                i, "Esfuerzo MPa M2", "Resistencia nominal"),
            "Esfuerzo % M3": lambda i: self.esfuerzo_pct_calc(
                i, "Esfuerzo MPa M3", "Resistencia nominal"),
            "Esfuerzo MPa Promedio": lambda i: self.promedio_calc(
                i, ["Esfuerzo MPa M1", "Esfuerzo MPa M2", "Esfuerzo MPa M3"]),
            "Esfuerzo % Promedio": lambda i: self.promedio_calc(
                i, ["Esfuerzo % M1", "Esfuerzo % M2", "Esfuerzo % M3"]),
        }

        # FIX #6: desconectar señal durante la carga
        self.table.itemChanged.disconnect(self.on_cell_changed)

        self.table.setRowCount(len(self.df))
        self.table.setColumnCount(len(self.visible_columns))
        self.table.setHorizontalHeaderLabels(self.visible_columns)

        for i in range(len(self.df)):
            for j, col in enumerate(self.visible_columns):
                value = self._get_display_value(i, col)
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item.setForeground(QColor("#2c3e50"))  # color base, sobrescrito por _apply_row_colors

                # FIX #5: "Esfuerzo MPa Promedio" es el objetivo del optimizador,
                # se deja editable de forma explícita y marcada visualmente
                if col in self.calculated_columns and col != "Esfuerzo MPa Promedio":
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    item.setBackground(Qt.GlobalColor.lightGray)
                    item.setToolTip("🔹 Calculado automáticamente")
                elif col == "Esfuerzo MPa Promedio":
                    item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                    item.setToolTip("🎯 Edita este valor para optimizar cargas automáticamente")
                else:
                    item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                    item.setToolTip("✏️ Doble clic para editar")

                self.table.setItem(i, j, item)

        # Reconectar señal
        self.table.itemChanged.connect(self.on_cell_changed)

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.resizeColumnsToContents()
        # Aplicar colores condicionales a todas las filas
        for i in range(len(self.df)):
            self._apply_row_colors(i)
        self.table.selectRow(0)

    # ─────────────────────────────────────────────
    # Optimizador
    # ─────────────────────────────────────────────
    def optimize_cargas_for_esfuerzo_promedio(self, fila, esfuerzo_promedio_deseado):
        try:
            cargas_disponibles = self._detectar_cargas_disponibles(fila)
            if not cargas_disponibles:
                return None

            areas = []
            for carga_idx in cargas_disponibles:
                cols_diam = self._get_diametros_carga(fila, carga_idx)
                d = self.promedio(fila, cols_diam)
                if pd.isna(d) or d <= 0:
                    return None
                areas.append((math.pi * (d ** 2)) / 4)  # mm²

            area_promedio = sum(areas) / len(areas)
            fuerza_promedio_necesaria = esfuerzo_promedio_deseado * area_promedio  # N
            carga_promedio_necesaria = fuerza_promedio_necesaria / 1000            # kg

            cargas_optimas = self._distribuir_cargas(
                carga_promedio_necesaria, areas, len(cargas_disponibles)
            )
            if cargas_optimas is None:
                return None

            # Actualizar DataFrame y tabla
            self.table.itemChanged.disconnect(self.on_cell_changed)
            for i, carga_idx in enumerate(cargas_disponibles):
                col_carga = f"Carga {carga_idx}"
                nuevo_valor = round(cargas_optimas[i], 1)
                self.df.at[fila, col_carga] = nuevo_valor
                col_idx = self.visible_columns.index(col_carga)
                item = self.table.item(fila, col_idx)
                if item:
                    item.setText(str(nuevo_valor))
            self.table.itemChanged.connect(self.on_cell_changed)

            return cargas_optimas, cargas_disponibles

        except Exception as e:
            print(f"❌ Error optimizando fila {fila}: {e}")
            # Asegurar reconexión de señal si hubo excepción a mitad
            try:
                self.table.itemChanged.connect(self.on_cell_changed)
            except RuntimeError:
                pass
            return None

    def _detectar_cargas_disponibles(self, fila):
        cargas = []
        for i in [1, 2, 3]:
            col1, col2 = f"Diámetro {i}-1", f"Diámetro {i}-2"
            if (col1 in self.df.columns and col2 in self.df.columns
                    and pd.notna(self.df.iloc[fila][col1])
                    and pd.notna(self.df.iloc[fila][col2])
                    and self.df.iloc[fila][col1] > 0
                    and self.df.iloc[fila][col2] > 0):
                cargas.append(i)
        return cargas

    def _get_diametros_carga(self, fila, carga_idx):
        return [f"Diámetro {carga_idx}-1", f"Diámetro {carga_idx}-2"]

    # FIX #3: tolerancia relativa en lugar de absoluta
    def _distribuir_cargas(self, carga_promedio, areas, num_cargas):
        try:
            sum_areas = sum(areas)
            cargas_base = [
                carga_promedio * (area / sum_areas) * num_cargas
                for area in areas
            ]

            # Variación aleatoria ±10%
            cargas_variadas = [c * (1 + random.uniform(-0.10, 0.10)) for c in cargas_base]

            # Ajustar para mantener promedio exacto
            factor = (carga_promedio * num_cargas) / sum(cargas_variadas)
            cargas_ajustadas = [c * factor for c in cargas_variadas]

            # Reducir variación si supera 15%
            if num_cargas > 1:
                max_c, min_c = max(cargas_ajustadas), min(cargas_ajustadas)
                if (max_c - min_c) / carga_promedio > 0.15:
                    cargas_ajustadas = [
                        carga_promedio * (1 + random.uniform(-0.05, 0.05))
                        for _ in range(num_cargas)
                    ]
                    factor = (carga_promedio * num_cargas) / sum(cargas_ajustadas)
                    cargas_ajustadas = [c * factor for c in cargas_ajustadas]

            # FIX #3: tolerancia relativa (0.1% del valor)
            promedio_final = sum(cargas_ajustadas) / num_cargas
            tolerancia = max(0.01, abs(carga_promedio) * 0.001)
            if abs(promedio_final - carga_promedio) > tolerancia:
                return None

            return [round(c, 1) for c in cargas_ajustadas]
        except Exception:
            return None

    # ─────────────────────────────────────────────
    # Colores condicionales por fila
    # ─────────────────────────────────────────────
    def _apply_row_colors(self, row: int):
        """
        Aplica colores condicionales a las celdas de una fila:
        - Esfuerzo % Promedio y Esfuerzo MPa Promedio: rojo si fuera del rango
          según la edad de falla (EDAD).
        - Densidad M1/M2/M3: rojo si < 1800 o > 2400.
        """
        if self.df is None:
            return

        ROJO  = QColor("#e74c3c")
        NEGRO = QColor("#2c3e50")

        # ── Leer EDAD y Esfuerzo % Promedio ───────────────────────────────
        try:
            edad = float(self.df.iloc[row]["EDAD"])
        except (ValueError, TypeError):
            edad = None

        try:
            col_idx_pct = self.visible_columns.index("Esfuerzo % Promedio")
            item_pct = self.table.item(row, col_idx_pct)
            esfuerzo_pct = float(item_pct.text()) if item_pct and item_pct.text() else None
        except (ValueError, TypeError, ValueError):
            esfuerzo_pct = None

        # ── Determinar si Esfuerzo % está fuera del rango según edad ──────
        esfuerzo_rojo = False
        if edad is not None and esfuerzo_pct is not None:
            if edad <= 10:
                esfuerzo_rojo = esfuerzo_pct < 55 or esfuerzo_pct > 80
            elif edad <= 20:
                esfuerzo_rojo = esfuerzo_pct < 70 or esfuerzo_pct > 100
            else:
                esfuerzo_rojo = esfuerzo_pct < 100 or esfuerzo_pct > 130

        # ── Colorear Esfuerzo % Promedio y Esfuerzo MPa Promedio ──────────
        for col_name in ("Esfuerzo % Promedio", "Esfuerzo MPa Promedio"):
            if col_name not in self.visible_columns:
                continue
            j = self.visible_columns.index(col_name)
            item = self.table.item(row, j)
            if item:
                item.setForeground(ROJO if esfuerzo_rojo else NEGRO)

        # ── Colorear densidades ────────────────────────────────────────────
        for col_name in ("Densidad M1", "Densidad M2", "Densidad M3"):
            if col_name not in self.visible_columns:
                continue
            j = self.visible_columns.index(col_name)
            item = self.table.item(row, j)
            if not item or not item.text():
                continue
            try:
                densidad = float(item.text())
                fuera = densidad < 1800 or densidad > 2400
                item.setForeground(ROJO if fuera else NEGRO)
            except (ValueError, TypeError):
                pass


    # ─────────────────────────────────────────────
    # Cálculos
    # ─────────────────────────────────────────────
    def _get_display_value(self, fila, col):
        try:
            if col in self.calculated_columns:
                value = self.calculated_columns[col](fila)
            else:
                value = self.df.iloc[fila][col]
            if pd.isna(value) or value == "":
                return ""
            elif col == "EDAD":
                return str(int(value))
            elif any(x in col for x in ["Densidad", "Diámetro", "Esfuerzo"]):
                return f"{float(value):.2f}"
            else:
                return str(value)
        except Exception:
            return ""

    def promedio(self, fila, columnas):
        valores = [
            pd.to_numeric(self.df.iloc[fila][col], errors="coerce")
            for col in columnas if col in self.df.columns
        ]
        valores = [v for v in valores if pd.notna(v)]
        return sum(valores) / len(valores) if valores else float("nan")

    def densidad(self, fila, col_masa, cols_diametro, cols_longitud):
        d = self.promedio(fila, cols_diametro)
        l = self.promedio(fila, cols_longitud)
        m = pd.to_numeric(self.df.iloc[fila][col_masa], errors="coerce")
        if pd.isna(d) or pd.isna(l) or pd.isna(m) or d <= 0 or l <= 0:
            return float("nan")
        d_cm, l_cm = d / 10, l / 10
        volumen = (math.pi * (d_cm ** 2) / 4) * l_cm
        return (m / volumen) * 1000

    def esfuerzo(self, fila, col_carga, cols_diametro):
        d = self.promedio(fila, cols_diametro)
        f = pd.to_numeric(self.df.iloc[fila][col_carga], errors="coerce")
        if pd.isna(d) or pd.isna(f) or d <= 0:
            return float("nan")
        area = (math.pi * (d ** 2)) / 4
        return (f * 1000) / area

    def esfuerzo_pct_calc(self, fila, col_esfuerzo, col_resistencia):
        e = self.calculated_columns[col_esfuerzo](fila)
        r = pd.to_numeric(self.df.iloc[fila][col_resistencia], errors="coerce")
        return (e / r) * 100 if pd.notna(e) and pd.notna(r) and r > 0 else float("nan")

    def promedio_calc(self, fila, columnas):
        valores = []
        for col in columnas:
            val = (self.calculated_columns[col](fila)
                   if col in self.calculated_columns
                   else self.df.iloc[fila][col])
            val = pd.to_numeric(val, errors="coerce")
            if pd.notna(val):
                valores.append(val)
        return sum(valores) / len(valores) if valores else float("nan")

    # ─────────────────────────────────────────────
    # Eventos de tabla
    # ─────────────────────────────────────────────
    def keyPressEvent(self, event):
        """Captura Ctrl+Z y Ctrl+Y a nivel de ventana principal."""
        if event.modifiers() == Qt.KeyboardModifier.ControlModifier:
            if event.key() == Qt.Key.Key_Z:
                self.undo()
                return
            if event.key() == Qt.Key.Key_Y:
                self.redo()
                return
        super().keyPressEvent(event)

    def on_row_selected(self, item=None):
        row = self.table.currentRow()
        if row >= 0:
            self.selected_row = row
            self.lbl_row.setText(f"Fila: {row + 1}")

    def on_cell_changed(self, item):
        if self.is_updating or self.df is None:
            return

        row = item.row()
        col_idx = item.column()
        col_name = self.visible_columns[col_idx]

        if col_name in self.calculated_columns and col_name != "Esfuerzo MPa Promedio":
            return

        try:
            text = item.text().strip()
            new_value = pd.to_numeric(text, errors="coerce") if text else None
            old_value = self.df.at[row, col_name]
            self.df.at[row, col_name] = new_value
            # Registrar en undo stack solo si el valor realmente cambió
            if str(old_value) != str(new_value):
                self._undo_stack.append((row, col_name, old_value, new_value))
                if len(self._undo_stack) > self._max_undo:
                    self._undo_stack.pop(0)
                self._redo_stack.clear()  # nueva edición invalida el redo
        except Exception:
            pass

        # Si el usuario editó Esfuerzo MPa Promedio → optimizar cargas automáticamente
        if col_name == "Esfuerzo MPa Promedio":
            self.selected_row = row
            try:
                esfuerzo_deseado = float(item.text().strip())
                if esfuerzo_deseado > 0:
                    self._set_status("loading", "🎯 Optimizando...")
                    QApplication.processEvents()
                    resultado = self.optimize_cargas_for_esfuerzo_promedio(row, esfuerzo_deseado)
                    if resultado:
                        cargas_optimas, cargas_usadas = resultado
                        detalle = ", ".join(
                            [f"{cargas_usadas[i]}:{c}" for i, c in enumerate(cargas_optimas)]
                        )
                        self._set_status("ok", f"✅ {esfuerzo_deseado:.1f} MPa | Cargas: {detalle}")
                        self._full_recalculate_row()
                    else:
                        self._set_status("error", "❌ Sin diámetros válidos para optimizar")
            except (ValueError, AttributeError):
                pass
            return

        self.selected_row = row
        self.on_form_changed(col_name)

    def _full_recalculate_row(self, skip_esfuerzo_promedio=False):
        if self.is_updating or self.df is None:
            return

        row = self.selected_row
        self._set_status("loading", "🔄 Recalculando...")
        QApplication.processEvents()

        self.is_updating = True
        # Desconectar durante el recálculo para no disparar on_cell_changed
        self.table.itemChanged.disconnect(self.on_cell_changed)
        try:
            for j, col in enumerate(self.visible_columns):
                if col not in self.calculated_columns:
                    continue
                if skip_esfuerzo_promedio and col == "Esfuerzo MPa Promedio":
                    continue
                value = self.calculated_columns[col](row)
                display = f"{float(value):.2f}" if pd.notna(value) else ""
                item = self.table.item(row, j)
                if item:
                    item.setText(display)
            self._apply_row_colors(row)
            self._set_status("ok", "🟢 Listo")
        except Exception as e:
            self._set_status("error", f"❌ Error: {e}")
        finally:
            self.is_updating = False
            self.table.itemChanged.connect(self.on_cell_changed)


    # ─────────────────────────────────────────────
    # Deshacer / Rehacer  (Ctrl+Z / Ctrl+Y)
    # ─────────────────────────────────────────────
    def undo(self):
        if not self._undo_stack:
            self._set_status("error", "⚠️ Nada que deshacer")
            return
        row, col_name, old_value, new_value = self._undo_stack.pop()
        self._redo_stack.append((row, col_name, old_value, new_value))
        self._apply_cell_value(row, col_name, old_value)
        self._set_status("ok", f"↩️ Deshecho: {col_name} fila {row+1}")

    def redo(self):
        if not self._redo_stack:
            self._set_status("error", "⚠️ Nada que rehacer")
            return
        row, col_name, old_value, new_value = self._redo_stack.pop()
        self._undo_stack.append((row, col_name, old_value, new_value))
        self._apply_cell_value(row, col_name, new_value)
        self._set_status("ok", f"↪️ Rehecho: {col_name} fila {row+1}")

    def _apply_cell_value(self, row, col_name, value):
        """Aplica un valor al DataFrame y a la celda de la tabla sin disparar undo."""
        self.df.at[row, col_name] = value
        self.selected_row = row

        if col_name not in self.visible_columns:
            return

        self.table.itemChanged.disconnect(self.on_cell_changed)
        try:
            col_idx = self.visible_columns.index(col_name)
            item = self.table.item(row, col_idx)
            if item:
                display = "" if pd.isna(value) or value is None else str(value)
                item.setText(display)
            self.table.selectRow(row)
        finally:
            self.table.itemChanged.connect(self.on_cell_changed)

        # Si la columna es Esfuerzo MPa Promedio, disparar el optimizador
        # con el valor restaurado (igual que cuando el usuario lo edita manualmente)
        if col_name == "Esfuerzo MPa Promedio":
            try:
                esfuerzo_deseado = float(value)
                if esfuerzo_deseado > 0:
                    self._set_status("loading", "🎯 Optimizando...")
                    QApplication.processEvents()
                    resultado = self.optimize_cargas_for_esfuerzo_promedio(row, esfuerzo_deseado)
                    if resultado:
                        cargas_optimas, cargas_usadas = resultado
                        detalle = ", ".join(
                            [f"{cargas_usadas[i]}:{c}" for i, c in enumerate(cargas_optimas)]
                        )
                        self._set_status("ok", f"✅ {esfuerzo_deseado:.1f} MPa | Cargas: {detalle}")
                        self._full_recalculate_row(skip_esfuerzo_promedio=True)
                    else:
                        self._set_status("error", "❌ Sin diámetros válidos para optimizar")
            except (ValueError, TypeError):
                pass
        else:
            self._full_recalculate_row()

    def on_form_changed(self, col_name):
        self.change_timer.stop()
        self.change_timer.start(150)

    # ─────────────────────────────────────────────
    # Guardado
    # FIX #4: conservar tipos numéricos en Excel
    # ─────────────────────────────────────────────

    # ─────────────────────────────────────────────
    # Pestaña Pendientes — helpers de estado
    # ─────────────────────────────────────────────
    def _set_status_p(self, estado: str, texto: str):
        colores = {
            "ok":      ("color:#27ae60;", "background:rgba(39,174,96,0.15);"),
            "error":   ("color:#e74c3c;", "background:rgba(231,76,60,0.15);"),
            "loading": ("color:#f39c12;", "background:rgba(243,156,18,0.15);"),
        }
        c, bg = colores.get(estado, colores["ok"])
        self.lbl_status_p.setStyleSheet(
            f"font-size:14px; padding:8px; border-radius:6px; {c} {bg}"
        )
        self.lbl_status_p.setText(texto)

    def _on_tab_changed(self, index: int):
        if index == 1 and self.df_p is None:
            self.load_pendientes()

    # ─────────────────────────────────────────────
    # Pestaña Pendientes — carga
    # ─────────────────────────────────────────────
    def load_pendientes(self):
        if self.current_file == "":
            self._set_status_p("error", "❌ Carga el archivo principal primero")
            return
        try:
            self._set_status_p("loading", "⏳ Cargando pendientes...")
            QApplication.processEvents()
            self.df_p = pd.read_excel(self.current_file, sheet_name="Pendientes_Generar")

            
            if "Diametro" in self.df_p.columns:
                self.df_p["Diametro"] = pd.to_numeric(
                    self.df_p["Diametro"], errors="coerce"
                ).astype("Int64")
            if "Cantidad" in self.df_p.columns:
                self.df_p["Cantidad"] = pd.to_numeric(
                    self.df_p["Cantidad"], errors="coerce"
                ).fillna(1).astype(int)

            cols_medicion = [
                "Esfuerzo MPa Promedio",
                "Diámetro 1-1","Diámetro 1-2","Longitud 1-1","Longitud 1-2","Longitud 1-3","Masa 1","Carga 1",
                "Diámetro 2-1","Diámetro 2-2","Longitud 2-1","Longitud 2-2","Longitud 2-3","Masa 2","Carga 2",
                "Diámetro 3-1","Diámetro 3-2","Longitud 3-1","Longitud 3-2","Longitud 3-3","Masa 3","Carga 3", 
            ]
            
            
            for col in cols_medicion:
                if col not in self.df_p.columns:
                    self.df_p[col] = None
                    
            

            self.populate_table_p()
            self._set_status_p("ok", f"✅ {len(self.df_p)} pendientes cargados")
            self.status_bar.showMessage(f"📋 Pendientes: {len(self.df_p)} filas")
            
        except Exception as e:
            self._set_status_p("error", f"❌ {e}")
            print(f"❌ load_pendientes: {e}")

    # ─────────────────────────────────────────────
    # Pestaña Pendientes — poblar tabla
    # ─────────────────────────────────────────────
    def populate_table_p(self):
        if self.df_p is None:
            return


        cols_info = ["consecutivo","idext","Resistencia nominal","codobraconf","Diametro","Cantidad", "EDAD"]
        cols_medicion = [
            "Esfuerzo MPa Promedio",
            "Diámetro 1-1","Diámetro 1-2","Longitud 1-1","Longitud 1-2","Longitud 1-3","Masa 1","Carga 1",
            "Diámetro 2-1","Diámetro 2-2","Longitud 2-1","Longitud 2-2","Longitud 2-3","Masa 2","Carga 2",
            "Diámetro 3-1","Diámetro 3-2","Longitud 3-1","Longitud 3-2","Longitud 3-3","Masa 3","Carga 3",
        ]
        
        # Recalcular EDAD en Python (date2 - date1 en días)
        # Equivale a la fórmula Excel: =SI(J2=0,"-", J2-I2)
        if "date1" in self.df_p.columns and "date2" in self.df_p.columns:
            d1 = pd.to_datetime(self.df_p["date1"], errors="coerce")
            d2 = pd.to_datetime(self.df_p["date2"], errors="coerce")
            # Resultado: entero de días; NaN si alguna fecha falta
            self.df_p["EDAD"] = (d2 - d1).dt.days
                    
        self.visible_columns_p = [c for c in cols_info + cols_medicion if c in self.df_p.columns]
        self.readonly_cols_p   = set(cols_info)
        self.generated_cols_p  = set(cols_medicion) - {"Esfuerzo MPa Promedio"}

        self.table_p.itemChanged.disconnect(self.on_cell_changed_p)
        self.table_p.setRowCount(len(self.df_p))
        self.table_p.setColumnCount(len(self.visible_columns_p))
        self.table_p.setHorizontalHeaderLabels(self.visible_columns_p)

        for i in range(len(self.df_p)):
            for j, col in enumerate(self.visible_columns_p):
                val = self.df_p.iloc[i][col]
                display = self._fmt_p(col, val)
                item = QTableWidgetItem(display)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item.setForeground(QColor("#2c3e50"))

                if col in self.readonly_cols_p:
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    item.setBackground(Qt.GlobalColor.lightGray)
                    item.setToolTip("🔒 Solo lectura")
                elif col == "Esfuerzo MPa Promedio":
                    item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                    item.setBackground(Qt.GlobalColor.white)
                    item.setToolTip("🎯 Ingresa el esfuerzo MPa objetivo para generar datos")
                else:
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    item.setBackground(QColor("#eaf4fb"))
                    item.setToolTip("⚙️ Generado automáticamente")

                self.table_p.setItem(i, j, item)

        self.table_p.itemChanged.connect(self.on_cell_changed_p)
        self.table_p.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table_p.resizeColumnsToContents()
        self.table_p.selectRow(0)

    def _fmt_p(self, col, val):
        try:
            if pd.isna(val) or val is None or val == "":
                return ""
        except Exception:
            pass
        if any(x in col for x in ["Diámetro", "Longitud", "Esfuerzo", "Masa", "Carga"]):
            try:
                return f"{float(val):.2f}"
            except Exception:
                pass
        return str(val)

    # ─────────────────────────────────────────────
    # Pestaña Pendientes — eventos
    # ─────────────────────────────────────────────
    def on_row_selected_p(self, item=None):
        row = self.table_p.currentRow()
        if row >= 0:
            self.lbl_row_p.setText(f"Fila: {row + 1}")

    def on_cell_changed_p(self, item):
        if self.is_updating_p or self.df_p is None:
            return
        row = item.row()
        col_name = self.visible_columns_p[item.column()]
        if col_name != "Esfuerzo MPa Promedio":
            return
        try:
            esfuerzo = float(item.text().strip())
            if esfuerzo <= 0:
                raise ValueError
        except (ValueError, AttributeError):
            return
        self._generar_fila_p(row, esfuerzo)

    # ─────────────────────────────────────────────
    # Pestaña Pendientes — generación de datos
    # ─────────────────────────────────────────────
    # Parámetros por diámetro nominal
    _PARAMS_P = {
        101: {"longitud_nominal": 201.0, "masa_min": 3300, "masa_max": 3600},
        151: {"longitud_nominal": 301.0, "masa_min": 12000, "masa_max": 13000},
    }

    def _submuestra(self, diam_nominal, esfuerzo_mpa):
        p = self._PARAMS_P[diam_nominal]
        v = 0.015   # ±1.5%
        d1 = round(diam_nominal * (1 + random.uniform(-v, v)), 2)
        d2 = round(diam_nominal * (1 + random.uniform(-v, v)), 2)
        l1 = round(p["longitud_nominal"] * (1 + random.uniform(-v, v)), 2)
        l2 = round(p["longitud_nominal"] * (1 + random.uniform(-v, v)), 2)
        l3 = round(p["longitud_nominal"] * (1 + random.uniform(-v, v)), 2)
        masa = random.randint(p["masa_min"], p["masa_max"])
        d_prom = (d1 + d2) / 2
        area = math.pi * (d_prom ** 2) / 4
        carga = round((esfuerzo_mpa * area) / 1000, 1)
        return {"d1":d1,"d2":d2,"l1":l1,"l2":l2,"l3":l3,"masa":masa,"carga":carga}

    def _generar_fila_p(self, row: int, esfuerzo_mpa: float):
        
        try:
            diam = int(self.df_p.iloc[row]["Diametro"])
            cantidad = int(self.df_p.iloc[row]["Cantidad"])
            
        except (ValueError, TypeError, KeyError):
            self._set_status_p("error", f"❌ Fila {row+1}: diámetro o cantidad inválidos")
            return
        if diam not in self._PARAMS_P:
            self._set_status_p("error", f"❌ Diámetro {diam} no soportado (use 101 o 151)")
            return

        n = min(cantidad, 3)
        submuestras = [self._submuestra(diam, esfuerzo_mpa) for _ in range(n)]

        campos = {
            1: ("Diámetro 1-1","Diámetro 1-2","Longitud 1-1","Longitud 1-2","Longitud 1-3","Masa 1","Carga 1"),
            2: ("Diámetro 2-1","Diámetro 2-2","Longitud 2-1","Longitud 2-2","Longitud 2-3","Masa 2","Carga 2"),
            3: ("Diámetro 3-1","Diámetro 3-2","Longitud 3-1","Longitud 3-2","Longitud 3-3","Masa 3","Carga 3"),
        }

        self.is_updating_p = True
        self.table_p.itemChanged.disconnect(self.on_cell_changed_p)
        try:
            for idx, sm in enumerate(submuestras, start=1):
                cols = campos[idx]
                valores = [sm["d1"],sm["d2"],sm["l1"],sm["l2"],sm["l3"],sm["masa"],sm["carga"]]
                for col, val in zip(cols, valores):
                    self.df_p.at[row, col] = val
                    if col in self.visible_columns_p:
                        j = self.visible_columns_p.index(col)
                        item = self.table_p.item(row, j)
                        if item:
                            item.setText(self._fmt_p(col, val))
            self.df_p.at[row, "Esfuerzo MPa Promedio"] = esfuerzo_mpa
            self._set_status_p("ok", f"✅ Fila {row+1}: {n} submuestras — {esfuerzo_mpa:.1f} MPa — ⌀{diam} mm")
            self.status_bar.showMessage(f"📋 Fila {row+1} generada — {n} submuestras")
        except Exception as e:
            self._set_status_p("error", f"❌ Error fila {row+1}: {e}")
        finally:
            self.is_updating_p = False
            self.table_p.itemChanged.connect(self.on_cell_changed_p)

    # ─────────────────────────────────────────────
    # Pestaña Pendientes — guardado
    # ─────────────────────────────────────────────
    def save_pendientes(self):
        if self.df_p is None or self.current_file == "":
            self._set_status_p("error", "❌ No hay datos cargados")
            return
        try:
            from openpyxl import load_workbook, Workbook as WB
            self._set_status_p("loading", "💾 Guardando...")
            QApplication.processEvents()

            output_path = Path(self.current_file).parent / "pendientes_final.xlsx"
            wb = load_workbook(output_path) if output_path.exists() else WB()
            if "Pendientes_final" in wb.sheetnames:
                del wb["Pendientes_final"]
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            ws = wb.create_sheet("Pendientes_final")

            for c_num, col_name in enumerate(self.df_p.columns, 1):
                ws.cell(row=1, column=c_num, value=str(col_name))
            for r_idx, (_, row_data) in enumerate(self.df_p.iterrows()):
                r_num = r_idx + 2
                for c_num, value in enumerate(row_data.values, 1):
                    try:
                        if pd.isna(value):
                            cell_value = None
                        else:
                            raise TypeError
                    except TypeError:
                        try:
                            fv = float(value)
                            cell_value = int(fv) if fv == int(fv) else fv
                        except (ValueError, TypeError):
                            cell_value = str(value)
                    ws.cell(row=r_num, column=c_num, value=cell_value)

            wb.save(output_path)
            
            # ── GENERAR INFORMES ─────────────────────────────
            self._set_status_p("loading", "📄 Generando informes...")
            QApplication.processEvents()

            file_path = Path(self.current_file)

            generados, errores = self.generar_informes(file_path, df=self.df_p)

            msg = f"💾 Guardado ✓ ({len(self.df_p)} filas) | 📄 {generados} informes"
            if errores:
                msg += f" | ⚠️ {errores} errores"

            self._set_status_p("ok", msg)
            self.status_bar.showMessage(
                f"✅ pendientes_final.xlsx — {generados} informes generados"
            )
            
            
            self._set_status_p("ok", f"💾 Guardado ✓ ({len(self.df_p)} filas)")
            self.status_bar.showMessage(f"✅ pendientes_final.xlsx guardado")
        except Exception as e:
            self._set_status_p("error", f"❌ Error: {e}")
            print(f"❌ save_pendientes: {e}")

    def save_all(self):
        if self.df is None or self.current_file == "":
            self._set_status("error", "❌ Carga datos primero")
            return

        try:
            import openpyxl
            from openpyxl import Workbook, load_workbook

            self._set_status("loading", "💾 Guardando...")
            QApplication.processEvents()

            file_path = Path(self.current_file)

            # Guardar Data_final en archivo separado para no bloquear prueba.xlsx
            output_path = file_path.parent / "data_final.xlsx"
            if output_path.exists():
                try:
                    wb = load_workbook(output_path)
                except Exception:
                    wb = Workbook()
            else:
                wb = Workbook()

            if "Data_final" in wb.sheetnames:
                del wb["Data_final"]
            # Eliminar hoja por defecto si el workbook es nuevo
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            ws = wb.create_sheet("Data_final")

            # Headers
            for col_num, col_name in enumerate(self.df.columns, 1):
                ws.cell(row=1, column=col_num, value=str(col_name))

            # FIX #4: intentar convertir a número antes de escribir
            for row_idx, (_, row_data) in enumerate(self.df.iterrows()):
                row_num = row_idx + 2
                for col_num, value in enumerate(row_data.values, 1):
                    if pd.isna(value):
                        cell_value = None          # celda vacía real en Excel
                    else:
                        try:
                            cell_value = float(value)
                            # Convertir a int si es un entero exacto (ej. 28, 3)
                            if cell_value == int(cell_value):
                                cell_value = int(cell_value)
                        except (ValueError, TypeError):
                            cell_value = str(value)
                    ws.cell(row=row_num, column=col_num, value=cell_value)

            wb.save(output_path)

            # Verificación rápida
            wb_check = load_workbook(output_path)
            ws_check = wb_check["Data_final"]
            filas_guardadas = ws_check.max_row - 1
            cols_guardadas = ws_check.max_column

            self._set_status("loading", f"📄 Generando informes...")
            QApplication.processEvents()

            generados, errores = self.generar_informes(file_path)

            msg = f"💾 Guardado ✓ ({filas_guardadas} filas) | 📄 {generados} informes"
            if errores:
                msg += f" | ⚠️ {errores} errores"
            self._set_status("ok", msg)
            self.status_bar.showMessage(
                f"✅ Data_final actualizada — {filas_guardadas} filas — {generados} informes generados"
            )

        except Exception as e:
            print(f"❌ ERROR save_all: {e}")
            self._set_status("error", f"❌ Error al guardar: {e}")

    # ─────────────────────────────────────────────
    # Generación de informes individuales
    # Basado en VBA invE410: llena hoja INV E 410-13
    # y guarda una copia del xlsx en Informes/ID-XXXXXXX-YYYY/
    # ─────────────────────────────────────────────
    def generar_informes(self, file_path: Path, df=None):
        """
        Por cada fila usa win32com para:
          1. Abrir prueba.xlsx en Excel (invisible)
          2. Llenar la hoja INV410INF con los datos de esa fila
          3. Copiar solo esa hoja a un libro nuevo (igual que VBA)
          4. Guardar el libro nuevo como xlsx en Informes/ID-XXXXXXX-YYYY/
          5. Exportar a PDF desde esa misma hoja
          6. Cerrar todo sin guardar cambios en prueba.xlsx
        """
        df = df if df is not None else self.df
        
        import sys as _sys
        from datetime import datetime

        if _sys.platform != "win32":
            print("⚠️ La generación de informes requiere Windows con Excel instalado.")
            return 0, len(self.df)

        try:
            import win32com.client
            import pythoncom
        except ImportError:
            print("⚠️ win32com no instalado. Ejecuta: pip install pywin32")
            return 0, len(self.df)

        anio     = datetime.now().year
        hoy_str  = datetime.now().strftime("%Y-%m-%d")
        informes_dir = file_path.parent / "Informes"
        informes_dir.mkdir(exist_ok=True)

        HOJA_INFORME = "INV410INF"

        # Mapeo celda → columna DataFrame  (tipo "fecha" formatea a YYYY-MM-DD)
        MAPA = {
            "BW1":  ("consecutivo",         None),
            "Q3":   ("campo23",             None),
            "G4":   ("fecharecep",          "fecha"),
            "Q4":   ("date2",               "fecha"),
            "G5":   ("cliente",             None),
            "G6":   ("dircliente",          None),
            "R6":   ("ciudadcliente",       None),
            "X6":   ("depcliente",          None),
            "AD6":  ("paiscliente",         None),
            "G7":   ("codobraconf",         None),
            "N7":   ("nameobra",            None),
            "G8":   ("procedencia",         None),
            "F10":  ("date1",               "fecha"),
            "O10":  ("Resistencia nominal", None),
            "Q12":  ("idext",               None),
            "Q13":  ("loc",                 None),
            "V10":  ("tipomat",             None),
            "AB4":  ("_fecha_hoy",          "fecha"),
            "Q17":  ("Longitud 1-1",        None),
            "Q18":  ("Longitud 1-2",        None),
            "Q19":  ("Longitud 1-3",        None),
            "Q20":  ("Diámetro 1-1",        None),
            "Q21":  ("Diámetro 1-2",        None),
            "Q22":  ("Masa 1",              None),
            "Q25":  ("Carga 1",             None),
            "Q26":  ("inv410ag",            None),
            "V17":  ("Longitud 2-1",        None),
            "V18":  ("Longitud 2-2",        None),
            "V19":  ("Longitud 2-3",        None),
            "V20":  ("Diámetro 2-1",        None),
            "V21":  ("Diámetro 2-2",        None),
            "V22":  ("Masa 2",              None),
            "V25":  ("Carga 2",             None),
            "V26":  ("inv410an",            None),
            "AA17": ("Longitud 3-1",        None),
            "AA18": ("Longitud 3-2",        None),
            "AA19": ("Longitud 3-3",        None),
            "AA20": ("Diámetro 3-1",        None),
            "AA21": ("Diámetro 3-2",        None),
            "AA22": ("Masa 3",              None),
            "AA25": ("Carga 3",             None),
            "AA26": ("inv410au",            None),
            "G28":  ("equipos",             None),
        }

        generados = 0
        errores   = 0

        pythoncom.CoInitialize()
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible        = False
        excel.DisplayAlerts  = False
        excel.ScreenUpdating = False

        try:
            for fila_idx, (_, fila) in enumerate(df.iterrows()):
                try:
                    # ── Nombre carpeta / archivo ───────────────────────────
                    consecutivo = str(fila.get("consecutivo", "")).strip()
                    if not consecutivo:
                        continue
                    partes = consecutivo.split("-")
                    num    = partes[-1] if len(partes) > 1 else consecutivo
                    nombre = f"ID-{num}-{anio}"

                    carpeta = informes_dir / nombre
                    carpeta.mkdir(exist_ok=True)

                    xlsx_destino = carpeta / f"{nombre}.xlsx"
                    pdf_destino  = carpeta / f"{nombre}.pdf"

                    # ── Abrir prueba.xlsx ──────────────────────────────────
                    wb_plantilla = excel.Workbooks.Open(str(file_path.resolve()))

                    # Verificar hoja
                    nombres_hojas = [sh.Name for sh in wb_plantilla.Sheets]
                    if HOJA_INFORME not in nombres_hojas:
                        print(f"⚠️ Hoja '{HOJA_INFORME}' no encontrada")
                        wb_plantilla.Close(SaveChanges=False)
                        errores += 1
                        continue

                    hoja = wb_plantilla.Sheets(HOJA_INFORME)

                    # ── Llenar celdas (igual que VBA) ──────────────────────
                    for celda, (col_df, tipo) in MAPA.items():
                        if col_df == "_fecha_hoy":
                            valor = hoy_str
                        elif col_df not in fila.index:
                            continue
                        else:
                            valor = fila[col_df]
                            if pd.isna(valor):
                                valor = ""
                            elif tipo == "fecha":
                                try:
                                    valor = pd.Timestamp(valor).strftime("%Y-%m-%d")
                                except Exception:
                                    valor = str(valor)
                            else:
                                try:
                                    fv = float(valor)
                                    valor = int(fv) if fv == int(fv) else fv
                                except (ValueError, TypeError):
                                    valor = str(valor)
                        hoja.Range(celda).Value = valor

                    # ── Copiar solo la hoja a un libro nuevo (como VBA) ────
                    # After:=None + Before:=None crea libro nuevo con esa hoja
                    hoja.Copy()          # sin argumentos → libro nuevo
                    wb_nuevo = excel.ActiveWorkbook

                    # ── Guardar xlsx ───────────────────────────────────────
                    # 51 = xlOpenXMLWorkbook (.xlsx sin macros)
                    wb_nuevo.SaveAs(
                        str(xlsx_destino.resolve()),
                        FileFormat=51,
                    )
                    print(f"✅ xlsx: {xlsx_destino}")

                    # ── Exportar PDF ───────────────────────────────────────
                    wb_nuevo.Sheets(1).ExportAsFixedFormat(
                        Type=0,                              # xlTypePDF
                        Filename=str(pdf_destino.resolve()),
                        Quality=0,                           # xlQualityStandard
                        IncludeDocProperties=True,
                        IgnorePrintAreas=False,
                        OpenAfterPublish=False,
                    )
                    print(f"✅ pdf:  {pdf_destino}")

                    wb_nuevo.Close(SaveChanges=False)

                    # Cerrar plantilla SIN guardar cambios
                    wb_plantilla.Close(SaveChanges=False)

                    generados += 1

                except Exception as e:
                    errores += 1
                    print(f"❌ Error fila {fila_idx}: {e}")
                    try:
                        excel.DisplayAlerts = False
                        for wb in list(excel.Workbooks):
                            wb.Close(SaveChanges=False)
                    except Exception:
                        pass

        finally:
            excel.ScreenUpdating = True
            excel.Quit()
            pythoncom.CoUninitialize()

        return generados, errores
    
    def refresh_excel(self, file_path: Path):
        import sys as _sys

        if _sys.platform != "win32":
            print("⚠️ Refresh requiere Excel en Windows")
            return

        try:
            import win32com.client
            import pythoncom

            pythoncom.CoInitialize()

            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            wb = excel.Workbooks.Open(str(file_path.resolve()))

            # 🔥 CLAVE: RefreshAll
            wb.RefreshAll()

            # ⏳ Esperar a que termine (MUY IMPORTANTE)
            excel.CalculateUntilAsyncQueriesDone()

            wb.Save()
            wb.Close(SaveChanges=True)

            excel.Quit()
            pythoncom.CoUninitialize()

            print("✅ Excel actualizado correctamente")

        except Exception as e:
            print(f"❌ Error refrescando Excel: {e}")


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    # Sobreescribir el color de resaltado (Highlight) de la paleta Fusion
    # Fusion usa QPalette.Highlight para pintar el fondo de foco/selección,
    # ignorando el CSS — por eso aparece rojo en algunas configuraciones de Windows.
    palette = app.palette()
    palette.setColor(QPalette.ColorRole.Highlight,        QColor("#1e8449"))  # verde oscuro
    palette.setColor(QPalette.ColorRole.HighlightedText,  QColor("#ffffff"))
    app.setPalette(palette)

    app.setStyleSheet("""
        QMainWindow {
            background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
                stop:0 #2c3e50, stop:1 #34495e);
            color: #ecf0f1;
        }

        QGroupBox {
            font-weight: bold;
            font-size: 14px;
            border: 1px solid #27ae60;
            border-radius: 12px;
            margin: 15px;
            padding-top: 20px;
            background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
                stop:0 #ffffff, stop:1 #f8f9fa);
            color: #2c3e50;
        }
        QGroupBox::title {
            subcontrol-origin: margin;
            left: 20px;
            padding: 0 12px 5px 12px;
            color: white;
            background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                stop:0 #27ae60, stop:1 #2ecc71);
            border-radius: 8px;
            font-size: 13px;
        }

        QTableWidget {
            gridline-color: #dee2e6;
            background: white;
            font-size: 11px;
            font-family: 'Segoe UI', Arial;
            border: 1px solid #2c3e50;
            alternate-background-color: #f8f9fa;
        }
        QTableWidget::item:selected {
            background: #27ae60;
            font-weight: bold;
        }
        QTableWidget::item:focus {
            border: 2px solid #1e8449;
            background: rgba(30,132,73,0.25);
        }

        QHeaderView::section {
            background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                stop:0 #2c3e50, stop:1 #34495e);
            color: #ffffff;
            font-weight: bold;
            padding: 12px;
            border: none;
            font-size: 11px;
        }

        QPushButton {
            padding: 14px 28px;
            font-weight: bold;
            border-radius: 10px;
            border: 2px solid transparent;
            font-size: 14px;
            font-family: 'Segoe UI', Arial;
            min-width: 140px;
        }
        QPushButton#btn_save {
            background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
                stop:0 #2ecc71, stop:1 #58d68d);
            color: white;
            border-color: #27ae60;
        }
        QPushButton#btn_save:hover {
            background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
                stop:0 #27ae60, stop:1 #2ecc71);
        }
        QPushButton#btn_save:pressed { background: #1e8449; }

        QPushButton#btn_reload {
            background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
                stop:0 #95a5a6, stop:1 #bdc3c7);
            color: #2c3e50;
            border-color: #7f8c8d;
        }
        QPushButton#btn_reload:hover {
            background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
                stop:0 #7f8c8d, stop:1 #95a5a6);
        }

        QLabel {
            color: #2c3e50;
            font-weight: 600;
            padding: 8px;
            background: rgba(255,255,255,0.8);
            border-radius: 6px;
        }

        QStatusBar {
            background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                stop:0 #34495e, stop:1 #2c3e50);
            color: #ecf0f1;
            font-weight: 500;
            padding: 8px;
            border-top: 2px solid #27ae60;
        }

        QScrollBar:vertical {
            background: #ecf0f1;
            width: 12px;
            border-radius: 6px;
        }
        QScrollBar::handle:vertical {
            background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                stop:0 #27ae60, stop:1 #2ecc71);
            border-radius: 6px;
            min-height: 20px;
        }
    """)

    window = ExcelDynamicApp()

    # Asignar objectName para los selectores CSS por ID (más robusto que setProperty)
    window.btn_save.setObjectName("btn_save")
    window.btn_reload.setObjectName("btn_reload")

    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()